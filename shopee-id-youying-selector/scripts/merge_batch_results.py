#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_COLUMNS = [
    "产品名称",
    "产品链接",
    "站点",
    "类目",
    "月销量",
    "价格(IDR)",
    "评分",
    "是否Shopee Mall",
    "上架时间",
    "店铺名称",
    "筛选结果",
    "淘汰原因",
    "抓取日期",
]

MERGED_COLUMNS = ["批次开始日期", "批次结束日期", *BASE_COLUMNS]


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(len(value) for value in values) + 2, 80)
        ws.column_dimensions[column_cells[0].column_letter].width = width


def load_summary(path: Path) -> list[dict]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, list):
        raise SystemExit("Batch summary JSON must be an array.")
    return data


def load_blacklist(path: Path | None) -> dict:
    if path is None:
        return {"product_links": [], "shop_names": [], "keywords": []}
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise SystemExit("Blacklist JSON must be an object.")
    return {
        "product_links": data.get("product_links", []),
        "shop_names": data.get("shop_names", []),
        "keywords": data.get("keywords", []),
    }


def load_rows(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path)
    ws = wb["products"]
    header = [cell.value for cell in ws[1]]
    rows: list[dict] = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = {header[idx]: values[idx] for idx in range(len(header))}
        rows.append(row)
    return rows


def should_skip(row: dict, blacklist: dict) -> tuple[bool, str]:
    product_link = str(row.get("产品链接", "") or "")
    shop_name = str(row.get("店铺名称", "") or "")
    title = str(row.get("产品名称", "") or "")
    category = str(row.get("类目", "") or "")
    haystack = f"{title} {shop_name} {category}".lower()

    for blocked in blacklist["product_links"]:
        if blocked and product_link == blocked:
            return True, f"blacklist_product_link:{blocked}"
    for blocked in blacklist["shop_names"]:
        if blocked and shop_name.lower() == str(blocked).lower():
            return True, f"blacklist_shop_name:{blocked}"
    for blocked in blacklist["keywords"]:
        if blocked and str(blocked).lower() in haystack:
            return True, f"blacklist_keyword:{blocked}"
    return False, ""


def write_merged_workbook(summary_rows: list[dict], output_path: Path, blacklist: dict, dedupe_by: str) -> None:
    wb = Workbook()

    all_ws = wb.active
    all_ws.title = "all_products"
    all_ws.append(MERGED_COLUMNS)

    filtered_ws = wb.create_sheet("filtered_products")
    filtered_ws.append([*MERGED_COLUMNS, "去重键", "过滤原因"])

    summary_ws = wb.create_sheet("batch_summary")
    summary_ws.append(["开始日期", "结束日期", "数量", "Excel", "JSON"])

    excluded_ws = wb.create_sheet("excluded")
    excluded_ws.append([*MERGED_COLUMNS, "去重键", "过滤原因"])

    total_count = 0
    filtered_count = 0
    seen_keys: set[str] = set()
    for item in summary_rows:
        start_date = item["start_date"]
        end_date = item["end_date"]
        xlsx_path = Path(item["output"])
        json_path = item.get("json_output", "")
        count = item.get("count", 0)
        summary_ws.append([start_date, end_date, count, str(xlsx_path), str(json_path)])

        for row in load_rows(xlsx_path):
            merged_row = [start_date, end_date, *[row.get(column, "") for column in BASE_COLUMNS]]
            all_ws.append(merged_row)
            total_count += 1

            dedupe_key = str(row.get("产品链接", "") if dedupe_by == "product_link" else row.get("店铺名称", ""))
            skip, reason = should_skip(row, blacklist)
            if not skip and dedupe_key and dedupe_key in seen_keys:
                skip, reason = True, f"duplicate_{dedupe_by}:{dedupe_key}"

            if skip:
                excluded_ws.append([*merged_row, dedupe_key, reason])
                continue

            if dedupe_key:
                seen_keys.add(dedupe_key)
            filtered_ws.append([*merged_row, dedupe_key, ""])
            filtered_count += 1

    meta_ws = wb.create_sheet("merge_meta")
    meta_ws.append(["字段", "值"])
    meta_ws.append(["批次数量", len(summary_rows)])
    meta_ws.append(["总商品数", total_count])
    meta_ws.append(["过滤后商品数", filtered_count])
    meta_ws.append(["去重方式", dedupe_by])
    meta_ws.append(["黑名单文件", json.dumps(blacklist, ensure_ascii=False)])
    meta_ws.append(["输出文件", str(output_path)])

    autosize_columns(all_ws)
    autosize_columns(filtered_ws)
    autosize_columns(summary_ws)
    autosize_columns(excluded_ws)
    autosize_columns(meta_ws)
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(description="Merge multiple batch Excel results into one workbook.")
    parser.add_argument("--summary-json", required=True, help="Batch summary JSON produced by batch_select_products.py.")
    parser.add_argument("--output", required=True, help="Output merged workbook path.")
    parser.add_argument("--blacklist-json", help="Optional blacklist JSON with product_links, shop_names, and keywords arrays.")
    parser.add_argument(
        "--dedupe-by",
        choices=["product_link", "shop_name"],
        default="product_link",
        help="How to deduplicate merged results.",
    )
    args = parser.parse_args()

    summary_rows = load_summary(Path(args.summary_json))
    blacklist = load_blacklist(Path(args.blacklist_json) if args.blacklist_json else None)
    write_merged_workbook(summary_rows, Path(args.output), blacklist, args.dedupe_by)
    print(json.dumps({"output": args.output, "runs": len(summary_rows), "dedupe_by": args.dedupe_by}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
